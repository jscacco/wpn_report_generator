import openpyxl
import datetime
import pickle

def read_wotc_skus():
        """Reads in the dict of wpn skus."""

        # I hate how gross and hacky this is, but I only want to read from this
        # file once the entire time, while also being able to consult new SKUS
        # entered for other transactions.
        
        # https://www.geeksforgeeks.org/how-to-read-dictionary-from-file-in-python/
        with open(DICT_FILENAME, 'rb') as handle:
            data = handle.read()
        return pickle.loads(data)


LG_ORG_ID = 40658
DG_ORG_ID = 35657
CURRENCY = "USD"
DICT_FILENAME = "files/wotc_sku_dict.txt"
WOTC_SKUS = read_wotc_skus()
NEW_SKUS = {}

class Transaction:
    def __init__(self, store):
        self.store = store
        self.wpn_org_id = None
        self.date = None
        self.transaction_id = None
        self.quantity_sold = None
        self.unit_price = None
        self.total_sale_price = None
        self.wotc_sku = None
        self.fg_product_desc = None
        self.currency = CURRENCY
        self.customer = None


    def __str__(self):
        output = ""
        output += "\n----------------------------------------\n"
        output += "Transaction " + str(self.transaction_id) + ": "
        output += "\n\n    WPN Org. ID: " +  str(self.wpn_org_id)
        output += "\n    wpn_org_id type: " +  str(type(self.wpn_org_id))
        output += "\n\n    Date: " + str(self.date)
        output += "\n    self.date type: " + str(type(self.date))
        output += "\n\n    Transaction ID: " + str(self.transaction_id)
        output += "\n    self.transaction_id type: " + str(type(self.transaction_id))
        output += "\n\n    Quantity Sold: " + str(self.quantity_sold)
        output += "\n    self.quantity_sold type: " + str(type(self.quantity_sold))
        output += "\n\n    Unit Price: " + str(self.unit_price)
        output += "\n    self.unit_price type: " + str(type(self.unit_price))
        output += "\n\n    Total Sale Price: " + str(self.total_sale_price)
        output += "\n    self.total_sale_price type: " +  str(type(self.total_sale_price))
        output += "\n\n    FG Product Desc: " + str(self.fg_product_desc)
        output += "\n    self.fg_product_desc type: " + str(type(self.fg_product_desc))
        output += "\n\n    WPN SKU: " + str(self.wotc_sku)
        output += "\n    self.wpn type: " + str(type(self.wotc_sku))
        output += "\n----------------------------------------\n"
        
        return output


    def getInfo(self, line_report_filename, r):
        """Given a line report filename and a row, read in the attributes.
           Note: This will not fill wotc SKU."""

        # Set wpn_org_id
        # This can be done whenever, but we do it here
        self.wpn_org_id = LG_ORG_ID if self.store == "LG" else DG_ORG_ID
        
        # Open the line report
        wb = openpyxl.load_workbook(line_report_filename)
        sheet = wb[wb.sheetnames[0]]
        
        for c in range(1, sheet.max_column + 1):
            this_value = sheet.cell(row=r, column=c).value

            if c == 1:
                # ID
                self.transaction_id = this_value
            elif c == 2:
                # Date
                self.date = this_value
            elif c == 3:
                # Description
                self.fg_product_desc = this_value
            elif c == 4:
                # Qty
                self.quantity_sold = this_value
            elif c == 5:
                # Retail (unit price)
                self.unit_price = this_value
            elif c == 6:
                # Subtotal
                self.total_sale_price = this_value
            elif c == 10:
                self.customer = this_value

                
    def adjustFormatDate(self):
        """If self.date is a string, convert it to a datetime object."""
        if type(self.date) == type('str'):
            split_str = self.date.split('-')
            self.date = datetime.datetime(int(split_str[0]), \
                                          int(split_str[1]), \
                                          int(split_str[2]))

                
    def adjustFormatPrices(self):
        """If self.unit_price is a string, convert it to a datetime object.
           Same for self.total_sale_price."""
        
        if type(self.unit_price) == type('str'):
            self.unit_price = float(remove_commas(self.unit_price[1:]))
            
        if type(self.total_sale_price) == type('str'):
            self.total_sale_price = float(remove_commas(self.total_sale_price[1:]))
            

    def getInfoAndFormat(self, line_report_filename, r):
        """ Consolidates the previous three methods."""
        self.getInfo(line_report_filename, r)
        self.adjustFormatDate()
        self.adjustFormatPrices()

        
    def isValidTransaction(self):
        """Returns True if the transaction should be added to the report."""

        # Basically, this catches all the transactions we don't want to add into
        # the report. This includes trade ins, admission fees, singles,
        # transaction with a total sale value not greater than zero,
        # and transactions made using the house account.
        
        desc = self.fg_product_desc.lower()
        return not (desc.__contains__("trade in") or \
                    desc.__contains__("admission") or \
                    desc.__contains__("single") or \
                    self.quantity_sold < 0)
                    
    
    def enterIntoWpnReport(self, wpn_report_filename, r):
        """Enter the values of the transaction into the WPN filename."""

        # Open the report
        wb = openpyxl.load_workbook(wpn_report_filename)
        sheet = wb[wb.sheetnames[0]]

        for c in range(1, sheet.max_column + 1):
            this_cell = sheet.cell(row=r, column=c)
            if c == 1:
                this_cell.value = self.wpn_org_id
            elif c == 2:
                this_cell.value = self.date
            elif c == 4:
                this_cell.value = self.transaction_id
            elif c == 6:
                this_cell.value = self.wotc_sku
            elif c == 9:
                this_cell.value = self.fg_product_desc
            elif c == 10:
                this_cell.value = self.quantity_sold
            elif c == 11:
                this_cell.value = self.unit_price
            elif c == 12:
                this_cell.value = self.total_sale_price
            elif c == 13:
                this_cell.value = self.currency

        wb.save(wpn_report_filename)

        
        
def get_line_report_col_descs(line_report_filename):
    """Given a line report, return a list of the column descriptions.
       Each description's column is equal to its index in the returned list.
       (Note that this means the list will always have filler info at its head)
       IMPORTANT: This assumes that the column descriptions will be in the first row.
       Update 7/29: I don't think this function is necessary."""

    # Open the line report
    col_descs = [None]
    wb = openpyxl.load_workbook(line_report_filename)
    sheet = wb[wb.sheetnames[0]]

    # Iterate through the first row, adding the column descriptions to the list
    for c in range(1, sheet.max_column + 1):
        col_descs.append(sheet.cell(row=1, column=c).value)

    # Return the list of descs
    return col_descs
    

def fill_wpn_report(store, line_report_filename, wpn_report_filename):
    """Put it all together."""
    wb = openpyxl.load_workbook(line_report_filename)
    sheet = wb[wb.sheetnames[0]]
    num_entries = sheet.max_row - 1

    current_line_row = 2
    current_wpn_row = 5

    while current_line_row <= num_entries:
        this_transaction = Transaction(store)
        this_transaction.getInfoAndFormat(line_report_filename, current_line_row)
        current_line_row += 1
        if this_transaction.isValidTransaction():
            set_wotc_sku(this_transaction)
            this_transaction.enterIntoWpnReport(wpn_report_filename,current_wpn_row)
            current_wpn_row += 1


def pickled_dict_setup(filenames):
    """Setup for the wpn SKU dict given a list of previous WPN reports.
       Should only be needed once."""

    this_dict = {}
    
    for report in filenames:
        print(report)
        wb = openpyxl.load_workbook(report)
        sheet = wb[wb.sheetnames[0]]

        # fg_desc = col 9
        # wpu_sku = col 6

        for r in range(5, sheet.max_row + 1):
            this_fg_desc = sheet.cell(row=r, column=9).value
            if this_fg_desc not in this_dict:
                this_dict[this_fg_desc] = sheet.cell(row=r, column=6).value

    file = open(DICT_FILENAME, 'wb')
    pickle.dump(this_dict, file)
    file.close()
    
    
def set_wotc_sku(transaction):
    """Gets and sets the wotc sku for the item."""

    if transaction.fg_product_desc == None:
        # I don't think this would ever trigger, but adding it just in case.
        print("\n\n\n!Error! Transaction has no product description.\n\n")
    elif transaction.fg_product_desc in WOTC_SKUS:
        transaction.wotc_sku = WOTC_SKUS[transaction.fg_product_desc]
    elif transaction.fg_product_desc in NEW_SKUS:
        transaction.wotc_sku = NEW_SKUS[transaction.fg_product_desc]
    else:
        # We need to input the SKU and add it in.
        confirmation = "n"
        while confirmation.lower() not in ['', 's']:
            new_sku = input("\nNo SKU found for " + transaction.fg_product_desc + \
                            ". Please enter it now: ")
            confirmation = input("Confirm: is " + new_sku + " the correct SKU for " + \
                                 transaction.fg_product_desc + \
                                 "? (Enter nothing if correct, S to skip, and any other button to re-enter the SKU)")

        if confirmation.lower() == 's':
                new_sku = "SKIPPED"
        NEW_SKUS[transaction.fg_product_desc] = new_sku
        transaction.wotc_sku = new_sku
        print("Working...")
        

def add_new_skus():
    """Adds the SKUS in NEW_SKUS to the dict file."""
    updated_skus = WOTC_SKUS
    updated_skus.update(NEW_SKUS)
    file = open(DICT_FILENAME, 'wb')
    pickle.dump(updated_skus, file)
    file.close()

    
def remove_commas(this_str):
        """Removes the commas from a given string."""
        new_str = ""
        for char in this_str:
                if char != ',':
                        new_str += char
        return new_str
        
    
def main():

    line_report_filename = "files/dg_0721_reports_sales_listings_transaction_line.xlsx"
    wpn_report_filename = "files/35657_FairGameDownersGrove_POSData_0721.xlsx"
    store = "DG"

    print("Working... (This will take a few minutes)")
    fill_wpn_report(store, line_report_filename, wpn_report_filename)
    add_new_skus()
                       
    # this_transaction = Transaction("DG")
    # this_transaction.getInfoAndFormat(line_report_filename, 2)
    # set_wotc_sku(this_transaction)
    # print(this_transaction)
    
if __name__ == "__main__":
    main()
