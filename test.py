import openpyxl

FILENAME = 'files/35657_FairGameDownersGrove_POSData_TEMPLATE.xlsx'

def part1():
    # Open the workbook and get the sheet names
    wb = openpyxl.load_workbook(FILENAME)
    sheet_names = wb.sheetnames
    # print(sheet_names)
    
    # Choose a sheet and access some of its values
    sheet = wb[sheet_names[0]]
    value = sheet['M5'].value
    print(value)
    print(type(value))

    # Make a change to one of the values and then save the sheet
    # b3_value = sheet['B3'].value
    # print(b3_value)
    # sheet['B3'].value = b3_value + 1
    # print(sheet['B3'].value)
    # wb.save(FILENAME)
    

def part2():
    # Open the workbook and desired sheet
    wb = openpyxl.load_workbook(FILENAME)
    sheet_names = wb.sheetnames
    sheet = wb[sheet_names[0]]

    # A new way to access cells. Supports iteration & loops!
    this_cell = sheet.cell(row=5, column=2)
    # print(type(this_cell))
    # print(this_cell.value)

    # Determine the range of cells which have values
    max_row = sheet.max_row
    max_col = sheet.max_column
    # print(max_row)
    # print(max_col)

    # How to convert from col str to index and vice versa
    letter = openpyxl.utils.get_column_letter(1)
    # print(letter)
    index = openpyxl.utils.column_index_from_string('AA')
    # print(index)

    # Loop through the values in the spreadsheet
    for c in range(1, max_col + 1):
        for r in range(1, max_row + 1):
            this_cell = sheet.cell(row=r, column=c)
            this_value = this_cell.value
            if r == 1:
                print(this_value + ": ")
            else:
                print(this_value)
        print()
            

def part3():
    # You already know
    wb = openpyxl.load_workbook(FILENAME)

    # Create a new sheet, let's use it
    wb.create_sheet(title="My Sheet", index=1)
    sheet = wb["My Sheet"]

    # Manually change the dimensions of a row and column
    sheet.row_dimensions[1].height = 70
    sheet.column_dimensions['B'].width = 50

    # Manually change the font of a single cell
    sheet['B1'].font = openpyxl.styles.Font(sz=14, bold=True, italic=True)
    sheet['B1'].value = "MACARONIIIIIIII"

    wb.save(FILENAME)

    
def main():
    part1()

    
if __name__ == "__main__":
    main()
