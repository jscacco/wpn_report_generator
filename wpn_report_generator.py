import openpyxl
import datetime
import pickle
import sys
import getopt
import math
import time

LG_ORG_ID = 40658
DG_ORG_ID = 35657
CURRENCY = "USD"

DICT_FILENAME = "files/wotc_sku_dict.txt"
ARR_FILENAME = "files/filter_keywords_array.txt"
NEW_SKUS = {}

DESC_COL_NUM = 3
QTY_COL_NUM = 4

class Transaction:
    def __init__(self, store):
        self.store = store
        self.wpn_org_id = None
        self.date = None
        self.transaction_id = None
        self.quantity_sold = None
        self.unit_price = None
        self.total_sale_price = None
        self.wotc_sku = None
        self.fg_product_desc = None
        self.currency = CURRENCY
        self.customer = None


    def __str__(self):
        output = ""
        output += "\n----------------------------------------\n"
        output += "Transaction " + str(self.transaction_id) + ": "
        output += "\n\n    WPN Org. ID: " +  str(self.wpn_org_id)
        output += "\n    wpn_org_id type: " +  str(type(self.wpn_org_id))
        output += "\n\n    Date: " + str(self.date)
        output += "\n    self.date type: " + str(type(self.date))
        output += "\n\n    Transaction ID: " + str(self.transaction_id)
        output += "\n    self.transaction_id type: " + str(type(self.transaction_id))
        output += "\n\n    Quantity Sold: " + str(self.quantity_sold)
        output += "\n    self.quantity_sold type: " + str(type(self.quantity_sold))
        output += "\n\n    Unit Price: " + str(self.unit_price)
        output += "\n    self.unit_price type: " + str(type(self.unit_price))
        output += "\n\n    Total Sale Price: " + str(self.total_sale_price)
        output += "\n    self.total_sale_price type: " +  str(type(self.total_sale_price))
        output += "\n\n    FG Product Desc: " + str(self.fg_product_desc)
        output += "\n    self.fg_product_desc type: " + str(type(self.fg_product_desc))
        output += "\n\n    WPN SKU: " + str(self.wotc_sku)
        output += "\n    self.wpn type: " + str(type(self.wotc_sku))
        output += "\n----------------------------------------\n"
        
        return output
    
    
    def getInfo(self, line_report_filename, r):
        """Given a line report filename and a row, read in the attributes.
           Note: This will not fill wotc SKU."""
        
        # Set wpn_org_id
        # This can be done whenever, but we do it here
        self.wpn_org_id = LG_ORG_ID if self.store == "LG" else DG_ORG_ID
        
        # Open the line report
        wb = openpyxl.load_workbook(line_report_filename)
        sheet = wb[wb.sheetnames[0]]

        self.transaction_id = sheet.cell(row=r, column=1).value
        self.date = sheet.cell(row=r, column=2).value
        self.fg_product_desc = sheet.cell(row=r, column=3).value
        self.quantity_sold = sheet.cell(row=r, column=4).value
        self.unit_price = sheet.cell(row=r, column=5).value
        self.total_sale_price = sheet.cell(row=r, column=6).value
        self.customere = sheet.cell(row=r, column=10).value

        wb.close()

        
    def adjustFormatDate(self):
        """If self.date is a string, convert it to a datetime object."""
        if type(self.date) == type('str'):
            split_str = self.date.split('-')
            self.date = datetime.datetime(int(split_str[0]), \
                                          int(split_str[1]), \
                                          int(split_str[2]))
            
            
    def adjustFormatPrices(self):
        """If self.unit_price is a string, convert it to a number.
           Same for self.total_sale_price."""
        
        if type(self.unit_price) == type('str'):
            self.unit_price = float(remove_commas(self.unit_price[1:]))
            
        if type(self.total_sale_price) == type('str'):
            self.total_sale_price = float(remove_commas(self.total_sale_price[1:]))
            

    def getInfoAndFormat(self, line_report_filename, r):
        """ Consolidates the previous three methods."""
        self.getInfo(line_report_filename, r)
        self.adjustFormatDate()
        self.adjustFormatPrices()

        
    def enterIntoWpnReport(self, wpn_report_filename, r):
        """Enter the values of the transaction into the WPN filename."""

        # Open the report
        wb = openpyxl.load_workbook(wpn_report_filename)
        sheet = wb[wb.sheetnames[0]]

        sheet.cell(row=r, column=1).value = self.wpn_org_id
        sheet.cell(row=r, column=2).value = self.date
        sheet.cell(row=r, column=4).value = self.transaction_id
        sheet.cell(row=r, column=6).value = self.wotc_sku
        sheet.cell(row=r, column=9).value = self.fg_product_desc
        sheet.cell(row=r, column=10).value = self.quantity_sold
        sheet.cell(row=r, column=11).value = self.unit_price
        sheet.cell(row=r, column=12).value = self.total_sale_price
        sheet.cell(row=r, column=13).value = self.currency

        wb.save(wpn_report_filename)
        wb.close()




def is_valid_transaction(line_report_filename, line_row):
    """Given a row in the line report, return true if that row contains a valid
       transaction."""

    wb = openpyxl.load_workbook(line_report_filename)
    sheet = wb[wb.sheetnames[0]]
    
    desc = sheet.cell(row=line_row, column=DESC_COL_NUM).value.lower()
    quantity_sold = sheet.cell(row=line_row, column=QTY_COL_NUM).value
    
    wb.close()

    keywords = fetch_filter_keywords()

    valid = True
    for k in keywords:
        if desc.__contains__(k):
            valid = False
            
    return valid and quantity_sold >= 0
    

def fill_wpn_report(store, line_report_filename, wpn_report_filename, wotc_skus):
    """Put it all together."""
    wb = openpyxl.load_workbook(line_report_filename)
    sheet = wb[wb.sheetnames[0]]
    num_entries = sheet.max_row - 1

    current_line_row = 2
    current_wpn_row = 5

    while current_line_row <= num_entries:
        if is_valid_transaction(line_report_filename, current_line_row):
            this_transaction = Transaction(store)
            this_transaction.getInfoAndFormat(line_report_filename, current_line_row)
            set_wotc_sku(this_transaction, wotc_skus)
            this_transaction.enterIntoWpnReport(wpn_report_filename, current_wpn_row)
            current_wpn_row += 1
        current_line_row += 1

    wb.close()


def pickled_dict_setup(filenames):
    """Setup for the wpn SKU dict given a list of previous WPN reports.
       Should only be needed once."""

    this_dict = {}
    
    for report in filenames:
        print(report)
        wb = openpyxl.load_workbook(report)
        sheet = wb[wb.sheetnames[0]]

        # fg_desc = col 9
        # wpu_sku = col 6

        for r in range(5, sheet.max_row + 1):
            this_fg_desc = sheet.cell(row=r, column=9).value
            if this_fg_desc not in this_dict:
                this_dict[this_fg_desc] = sheet.cell(row=r, column=6).value

        wb.close()

    file = open(DICT_FILENAME, 'wb')
    pickle.dump(this_dict, file)
    file.close()


        
def set_wotc_sku(transaction, wotc_skus):
    """Gets and sets the wotc sku for the item."""
    
    if transaction.fg_product_desc == None:
        # If there is not description...
        # (I don't think this would ever trigger, but adding it just in case.)
        print("\n\n\n!Error! Transaction has no product description.\n\n")
    elif transaction.fg_product_desc in wotc_skus:
        # If the SKU is already in our dictionary...
        transaction.wotc_sku = wotc_skus[transaction.fg_product_desc]
    elif transaction.fg_product_desc in NEW_SKUS:
        # If the SKU is new, but we've already seen it this session...
        transaction.wotc_sku = NEW_SKUS[transaction.fg_product_desc]
    else:
        # We need to input the SKU and add it in.
        confirmation = "n"
        while confirmation.lower() not in ['', 's']:
            new_sku = input("\nNo SKU found for " + transaction.fg_product_desc + \
                            ". Please enter it now: ")
            confirmation = input("Confirm: is " + new_sku + " the correct SKU for " + \
                                 transaction.fg_product_desc + \
                                 "? (Enter nothing if correct, S to skip, and any other button to re-enter the SKU)")

        if confirmation.lower() == 's':
            new_sku = "SKIPPED"
        NEW_SKUS[transaction.fg_product_desc] = new_sku
        transaction.wotc_sku = new_sku
        print("Working...")
        

def add_new_skus(wotc_skus):
    """Adds the SKUS in NEW_SKUS to the dict file."""
    updated_skus = wotc_skus
    updated_skus.update(NEW_SKUS)
    file = open(DICT_FILENAME, 'wb')
    pickle.dump(updated_skus, file)
    file.close()

    
def remove_commas(this_str):
    """Removes the commas from a given string."""
    new_str = ""
    for char in this_str:
        if char != ',':
            new_str += char
    return new_str


def seconds_to_formatted_time(seconds):
    """Given a number representing a length of time in seconds, return a string which
       formats the time into a more readable format."""

    seconds = int(seconds)
    
    h = math.floor(seconds / 3600)
    m = math.floor((seconds / 60) % 60)
    s = math.floor(seconds % 60)
    
    return "{h}h, {m}min, {s}sec".format(h=h, m=m, s=s)

    
def generate_report(line_report_filename, wpn_report_filename, store):
    """The bulk of the program, generate a new report."""
    

    # let's keep track of how long this takes
    start = time.time()
    wotc_skus = fetch_wotc_skus()
    
    print("Working... (This will take a few minutes)")
    fill_wpn_report(store, line_report_filename, wpn_report_filename, wotc_skus)
    add_new_skus(wotc_skus)

    end = time.time()
    
    print("Finished! Time elapsed: " + seconds_to_formatted_time(end - start))


def print_help_info():
    """Show how to use the program."""
    
    print("\nGeneral usage (for report generation):")
    print("\twpn_report_generator.py -l <line report filename> -w <wpn report filename> -s <store (DG or LG)>")
    print("\nOther parameters:")
    print("\t-u (to view/update/lookup WotC SKUs)")
    print("\t-k (to view/add/delete filter keywords)")
    

def fetch_wotc_skus():
    """Reads in the dict of wpn skus."""
    
    # https://www.geeksforgeeks.org/how-to-read-dictionary-from-file-in-python/
    with open(DICT_FILENAME, 'rb') as handle:
        data = handle.read()
    return pickle.loads(data)


def display_wotc_skus():
    """Prints out the filter keywords to the user."""

    skus = fetch_wotc_skus()
    i = 0
    for key in skus:
        print(i, key, skus[key])
        i += 1

        
def lookup_sku():
    """Looks up a sku."""
    # Get the name of the item to be updated from the user and display the current SKU
    wotc_skus = fetch_wotc_skus()
        
    desc = input("Please enter the name of the item whose SKU you would like to lookup. \n")
    try:
        current_sku = wotc_skus[desc]
    except KeyError:
        print("No item found with that description.")
        return

    print("The current SKU for " + desc + " is: " + str(current_sku) + ".")

    
def update_sku():
    """Update an item's SKU."""
    wotc_skus = fetch_wotc_skus()
        
    desc = input("Please enter the name of the item whose SKU you would like to update:\n")
    try:
        current_sku = wotc_skus[desc]
    except KeyError:
        print("No item found with that description")
        return

    print("The current SKU for " + desc + " is: " + str(current_sku) + ".")
    
    # Get the new SKU from the user
    confirm = 'n'
    while confirm not in ['y', '', 'yes']:
        new_sku = input("Please enter the new SKU:\n")
        confirm = input("Is " + str(new_sku) + " correct? (Enter 'y', 'yes', or nothing to confirm. Enter any other key to resubmit):\n")

    # Update the SKU dict
    wotc_skus[desc] = new_sku
    file = open(DICT_FILENAME, 'wb')
    pickle.dump(wotc_skus, file)
    file.close()


def sku_manager():
    """Helper function for various sku-related functions."""
    command = input("Would you like to view (v) skus, update (u) a sku, lookup (l) a specific sku, or exit (any other key)? Enter the respective letter:\n").lower()

    if command ==  "v":
        display_wotc_skus()
    elif command == "u":
        update_sku()
    elif command == "l":
        lookup_sku()
        

def fetch_filter_keywords():
    """Returns the list containing the filter keywords."""

    with open(ARR_FILENAME, 'rb') as handle:
        data = handle.read()
        
    return pickle.loads(data)
    
    
def display_filter_keywords():
    """Prints out the filter keywords to the user."""

    keywords = fetch_filter_keywords()
    for i in range(len(keywords)):
        print(i, keywords[i])


def write_filter_keywords(keywords):
    """Given a list of keywords, write that list to ARR_FILENAME."""

    file = open(ARR_FILENAME, 'wb')
    pickle.dump(keywords, file)
    file.close()


def delete_filter_keyword(index):
    """Given a keyword's index, remove it from the filter keywords file."""
    keys = fetch_filter_keywords()
    if index < 0 or index > (len(keys) - 1):
        print("invalid index")
        return
    del keys[index]
    write_filter_keywords(keys)


def add_filter_keyword(keyword):
    """Given a keyword, add it to the list of filter keywords."""
    keys = fetch_filter_keywords()
    if keyword not in keys:
        keys.append(keyword)
    write_filter_keywords(keys)


def keyword_manager():
    """Helper function to manage various filter keyword functions."""
    command = input("Would you like to view (v) keywords, delete (d) a keyword, add (a) a keyword, or exit (any other key)? Enter the respective letter:\n").lower()

    if command == "v":
        display_filter_keywords()
    elif command == "d":
        index = int(input("Please enter the index number of the keyword to delete (displayed to the left of the keyword):\n"))
        delete_filter_keyword(index)
    elif command == "a":
        keyword = input("Enter the keyword you would like to add to the filter keyword list. (Case insensitive):\n")
        add_filter_keyword(keyword)
        

def main():
    line_report_filename = ''
    wpn_report_filename = ''
    store = ''
    
    try:
        opts, args = getopt.getopt(sys.argv[1:], "hl:w:s:uk", ["line=", "wpn=", "store="])
    except:
        print_help_info()
        sys.exit(2)

    # If there are no arguments, print the help and exit
    if len(opts) == 0:
        print_help_info()
        sys.exit(2)

    # Otherwise, determine functionality based on arguments provided
    for opt, arg in opts:
        if opt == '-h':
            print_help_info()
            sys.exit()
        elif opt == '-k':
            keyword_manager()
            sys.exit()
        elif opt == '-u':
            sku_manager()
            sys.exit()
        elif opt in ['-l', '--line']:
            line_report_filename = arg
        elif opt in ['-w', '--wpn']:
            wpn_report_filename = arg
        elif opt in ['-s', '--store']:
            store = arg

    # If we are generating a report and don't have the three arguments we need,
    # inform the user that they need to provide them and exit.
    if line_report_filename == '' or wpn_report_filename == '' or store == '':
        print("Please include a line report filename, wpn report filename, and store.")
        sys.exit(2)

    generate_report(line_report_filename, wpn_report_filename, store)


    
if __name__ == "__main__":
    main()
